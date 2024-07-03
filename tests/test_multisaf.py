"""Test [S]imple[A]rchive[F]ile generation format required for DSpace delivery"""

import os
import shutil

from unittest.mock import DEFAULT
import pytest
import openpyxl
import lxml.etree as ET

from digiflow.digiflow_multisaf import create_saf


from digiflow import (
    export_data_from,
    map_contents,
    ExportMapping,
    DEFAULT_EXPORT_MAPPINGS,
    DigiFlowExportError,
    BUNDLE_PREVIEW as BP,
    BUNDLE_THUMBNAIL as BT,
)

import digiflow.validate as df_v

from .conftest import (
    TEST_RES
)


def _check_data(extr_dir, expected_saf_files, ddict):
    with open(extr_dir / "contents", "r", encoding="UTF-8") as file:
        contents = file.readlines()
    assert (len(contents)) == (len(os.listdir(extr_dir)) - 4)  # All expected files are available
    for contentline in contents:
        fn = contentline.split("\t")[0].split("\n")[0]
        assert fn in expected_saf_files[ddict["Folder"]]
        expected_saf_files[ddict["Folder"]].remove(fn)
    assert len(expected_saf_files[ddict["Folder"]]) == 0

    path_dublic_core = extr_dir / "dublin_core.xml"
    dc_tree = ET.parse(path_dublic_core).getroot()
    assert df_v.validate_xml(dc_tree)

    with open(path_dublic_core, "r", encoding="UTF-8") as file:
        dublin_core = file.read()
    for (element, qualifier, content) in ddict["Content"]:
        if qualifier != "":
            if len(qualifier.split("|||")) == 1:
                write_string = '  <dcvalue element="' + \
                    str(element) + '" qualifier="' + str(qualifier) + \
                    '">' + str(content) + '</dcvalue>\n'
            else:
                if len(content.split("|||")) > 1:
                    cont, auth = content.split("|||")
                    write_string = '  <dcvalue element="' + str(element) +\
                        '" qualifier="' + str(qualifier.split("|||")[0]) +\
                            '" authority="' + auth + '" confidence="' +\
                                str(600) + '">' + str(cont) + '</dcvalue>\n'
                else:
                    write_string = '  <dcvalue element="' + \
                        str(element) + '" qualifier="' + str(qualifier.split("|||")
                            [0]) + '">' + str(content) + '</dcvalue>\n'
        else:
            write_string = '  <dcvalue element="' + \
                str(element) + '">' + str(content) + '</dcvalue>\n'
        assert write_string in dublin_core  # All metadata seems to be correct


def _setup_data(cnt_dir, table_path):
    expected_saf_files = {}
    for to_saf in os.listdir(cnt_dir / "to_saf/"):
        expected_saf_files[to_saf] = []
        for (fullpath, _, files) in os.walk(cnt_dir / "to_saf/" / to_saf):
            rel_dir = os.path.relpath(fullpath, cnt_dir / "to_saf/" / to_saf)
            clean_rel_dir = ""
            for char in rel_dir:
                if not char.isalnum():
                    clean_rel_dir = clean_rel_dir + "_"
                else:
                    clean_rel_dir = clean_rel_dir + char
            for file in files:
                if file not in expected_saf_files[to_saf]:
                    expected_saf_files[to_saf].append(file)
                else:
                    expected_saf_files[to_saf].append(clean_rel_dir + "_" + file)

    # Get table data
    wb = openpyxl.load_workbook(table_path)
    metadata_table = wb.active
    metadata_table_meta = {}
    metadata_table_dc = {}
    for cell in metadata_table[2]:
        if cell.value:
            if cell.value.split(".")[0] == "meta":
                metadata_table_meta[cell.value.split(".")[-1]] = cell.column
            if cell.value.split(".")[0] == "dc":
                metadata_table_dc[cell.value] = cell.column

    data_dict = {}
    for row_number in range(3, metadata_table.max_row + 1):
        foldername = str(metadata_table.cell(
            row=row_number, column=metadata_table_meta["ordner"]).value)
        data_dict[row_number] = {}
        data_dict[row_number]["Folder"] = foldername
        data_dict[row_number]["Content"] = []
        for dcentry, td_dc in metadata_table_dc.items():
            split_dc = dcentry.split(".")
            if len(split_dc) == 3:
                (_, element, qualifier) = split_dc
            else:
                if len(split_dc) == 2:
                    split_dc.append("")
                    (_, element, qualifier) = split_dc
            content = metadata_table.cell(row=row_number, column=td_dc).value
            if content:
                if isinstance(content, str) and ";" in content:
                    content = content.split(";")
                    for sub_content in content:
                        data_dict[row_number]["Content"].append((element, qualifier, sub_content))
                else:
                    data_dict[row_number]["Content"].append((element, qualifier, content))
    return data_dict, expected_saf_files


@pytest.fixture(name="multisaf")
def fixture_multisaf(tmp_path):
    """fixture for multisaf"""
    src_dir = os.path.join(TEST_RES, 'multisaf')
    dst_dir = tmp_path / 'content'
    shutil.copytree(src_dir, dst_dir)
    work_dir = tmp_path / "export_workdir"
    work_dir.mkdir()
    exp_dir = tmp_path / "export_finished"
    exp_dir.mkdir()
    yield (dst_dir, work_dir, exp_dir)


def test_correct_contents(multisaf):
    """Check if the contents are the expected ones"""
    # arrange
    (cnt_dir, _, exp_dir) = multisaf
    # act
    one_saf = False
    table_path = cnt_dir / "Metadata_Collection.xlsx"
    create_saf(table_path, cnt_dir / "to_saf/", one_saf, exp_dir)

    # Find files to saf
    data_dict, expected_saf_files = _setup_data(cnt_dir, table_path)
    # Asserting
    for saf_file in os.listdir(exp_dir):
        foldername = saf_file.split(".")[0]
        extract_saf_dir = exp_dir / foldername
        shutil.unpack_archive(exp_dir / saf_file, extract_dir=extract_saf_dir)

    assert len(os.listdir(exp_dir))/2 == len(data_dict.items())
    for ddict in data_dict.values():
        assert ddict["Folder"] in os.listdir(exp_dir)
        extr_dir = exp_dir / ddict["Folder"] / ddict["Folder"]
        _check_data(extr_dir, expected_saf_files, ddict)


def test_correct_contents_onesaf(multisaf):
    """Check if the contents are the expected ones"""
    # arrange
    (cnt_dir, _, exp_dir) = multisaf
    # act
    one_saf = True
    table_path = cnt_dir / "Metadata_Collection.xlsx"
    create_saf(table_path, cnt_dir / "to_saf/", one_saf, exp_dir)

    # Find files to saf
    data_dict, expected_saf_files = _setup_data(cnt_dir, table_path)
    # Asserting
    for saf_file in os.listdir(exp_dir):
        foldername = saf_file.split(".")[0]
        extract_saf_dir = exp_dir / foldername
        shutil.unpack_archive(exp_dir / saf_file, extract_dir=extract_saf_dir)

    assert len(os.listdir(exp_dir))/2 == 1
    for item_folder in os.listdir(extract_saf_dir):
        row_num = int(item_folder.split("item_")[-1]) + 3
        extr_dir = extract_saf_dir / item_folder
        ddict = data_dict[row_num]
        _check_data(extr_dir, expected_saf_files, ddict)
