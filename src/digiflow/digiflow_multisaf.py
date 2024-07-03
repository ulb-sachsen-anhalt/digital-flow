"""Create multiple SAF-files from metadata table"""
import argparse
import os
import shutil
import pathlib

import openpyxl


def write_dc(element, qualifier, content, xmlfile):
    """Create DC file"""
    if qualifier != "":
        if len(qualifier.split("|||")) == 1:
            write_string = '  <dcvalue element="' + \
                str(element) + '" qualifier="' + str(qualifier) + \
                '">' + str(content) + '</dcvalue>\n'
        else:
            if len(content.split("|||")) > 1:
                cont, auth = content.split("|||")
                write_string = '  <dcvalue element="' + str(element) + '" qualifier="'\
                + str(qualifier.split("|||")[0]) + '" authority="'\
                + auth + '" confidence="' + str(600) + '">' + str(cont) + '</dcvalue>\n'
            else:
                write_string = '  <dcvalue element="' + \
                    str(element) + '" qualifier="' + str(qualifier.split("|||")
                                                         [0]) + '">' + str(content) + '</dcvalue>\n'
    else:
        write_string = '  <dcvalue element="' + str(element) + '">' + str(content) + '</dcvalue>\n'
    xmlfile.write(write_string)


def find_files(file_list, name_pre, path):
    """Check files and create proper name"""
    dirlist = []
    file_list_cur = []
    for diritem in os.listdir(path):
        if os.path.isdir(path / pathlib.Path(diritem)):
            dirlist.append(diritem)
        else:
            file_list_cur.append(diritem)
    for dire in dirlist:
        sdir = ""
        for char in str(dire):
            if char.isalnum():
                sdir = sdir + char
            else:
                sdir = sdir + "_"
        file_list = find_files(file_list, name_pre + sdir + "_", path / pathlib.Path(dire))
    for f in file_list_cur:
        file_list.append((path / pathlib.Path(f), (name_pre + str(f))))
    return file_list


def create_saf(table_path, file_path_base, one_saf_input, finished_folder):
    """Create the SAF file"""
    wb = openpyxl.load_workbook(table_path)
    if one_saf_input: # Making sure it is bool
        one_saf = True
    else:
        one_saf = False
    metadata_table = wb.active
    metadata_table_meta = {}
    metadata_table_dc = {}
    print("Starting preparation...")
    print("Reading the table")
    for cell in metadata_table[2]:
        if cell.value:
            if cell.value.split(".")[0] == "meta":
                metadata_table_meta[cell.value.split(".")[-1]] = cell.column
            if cell.value.split(".")[0] == "dc":
                metadata_table_dc[cell.value] = cell.column
    base_file_path = pathlib.Path(file_path_base)
    project_name = base_file_path.name
    finished_folder = base_file_path.parent / finished_folder
    print("Creating " + str(finished_folder))
    if not os.path.isdir(finished_folder):
        os.makedirs(finished_folder)
    print("Created " + str(finished_folder) + ", now starting actual saf creation...")
    item_number = 0
    number_of_items = int(metadata_table.max_row - 2)
    number_of_items_string_length = len(str(number_of_items))
    for row_number in range(3, metadata_table.max_row + 1):
        try:
            folder_name = str(metadata_table.cell(
                row=row_number, column=metadata_table_meta["ordner"]).value)
            file_path = base_file_path / folder_name
            print("Now working on " + str(folder_name) + " (Rownumber = " + str(row_number) + ")")
            if one_saf:
                temp_folder_base = base_file_path.parent / "One_Saf"
                item_number_string = str(item_number)
                if len(item_number_string) < number_of_items_string_length:
                    for _ in range(number_of_items_string_length - len(item_number_string)):
                        item_number_string = "0" + item_number_string
                temp_folder = temp_folder_base / pathlib.Path("item_" + item_number_string)
                item_number = item_number + 1
            else:
                temp_folder_base = base_file_path.parent / "TEMP"
                temp_folder = temp_folder_base / folder_name
            os.makedirs(temp_folder)
            # dublin_core.xml
            print("Creating dublin_core.xml")
            with open(temp_folder / "dublin_core.xml", "w", encoding="UTF-8") as xmlfile:
                xmlfile.write('<?xml version="1.0" encoding="UTF-8"?>\n')
                xmlfile.write('<dublin_core>\n')
                for dcentry, td_dc in metadata_table_dc.items():
                    split_dc = dcentry.split(".")
                    if len(split_dc) == 3:
                        (_, element, qualifier) = split_dc
                    else:
                        if len(split_dc) == 2:
                            split_dc.append("")
                            (_, element, qualifier) = split_dc
                    content = metadata_table.cell(row=row_number, column=td_dc).value
                    if content:
                        if isinstance(content, str) and ";" in content:
                            content = content.split(";")
                            for subcontent in content:
                                write_dc(element, qualifier, subcontent, xmlfile)
                        else:
                            write_dc(element, qualifier, content, xmlfile)

                xmlfile.write("</dublin_core>")

            # collections
            print("Creating collections")
            with open(temp_folder / "collections", "w", encoding="UTF-8") as file:
                file.write(
                    str(metadata_table.cell(row=row_number,\
                        column=metadata_table_meta["collection"]).value))
            # Filehandling
            print("Copying files to the right place")
            list_files_with_path = find_files([], "", file_path)
            double_file_check = []
            for (source_file, file_name) in list_files_with_path:
                shutil.copy(source_file, temp_folder / file_name)
                if file_name not in double_file_check:
                    double_file_check.append(file_name)
                else:
                    raise ValueError("Filename was used twice: " + str(file_name) + \
                                    "\nThis should only happen if there is a file" \
                                    "'dirname_file.txt', and a folder 'dirname'"\
                                    "with 'file.txt' in it")
            # contents
            print("Creating contents")
            rights_group = metadata_table.cell(
                row=row_number, column=metadata_table_meta["dspacegroup"]).value
            with open(temp_folder / "contents", "w", encoding="UTF-8") as file:
                if rights_group:
                    rights_group = str(rights_group)
                    for (_, file_name) in list_files_with_path:
                        file.write(file_name + "\tpermissions: -r '" + rights_group + "'\n")
                else:
                    for (_, file_name) in list_files_with_path:
                        file.write(file_name + "\n")
            # metadata_local.xml
            print("Creating metadata_local.xml")
            with open(temp_folder / "metadata_local.xml", "w", encoding="UTF-8") as xmlfile:
                xmlfile.write('<dublin_core schema="local">\n')
                xmlfile.write('  <dcvalue element="accessrights" qualifier="dnb">free</dcvalue>\n')
                if rights_group:
                    xmlfile.write('  <dcvalue element="openaccess" qualifier="">false</dcvalue>\n')
                else:
                    xmlfile.write('  <dcvalue element="openaccess" qualifier="">true</dcvalue>\n')
                xmlfile.write('</dublin_core>')
            # SAF-Creation

            if not one_saf:
                print("Finally creating the saf")
                if not folder_name.isalnum():
                    folder_name_clean = ""
                    for char in folder_name:
                        if not char.isalnum():
                            folder_name_clean = folder_name_clean + "_"
                        else:
                            folder_name_clean = folder_name_clean + char
                    folder_name = folder_name_clean
                shutil.make_archive(finished_folder / folder_name,
                                    format="zip", root_dir=temp_folder_base)
                shutil.rmtree(temp_folder_base)
                print("saf created")
            print("Finished " + str(folder_name) + " (Rownumber = " + str(row_number) + ")\n")
        except Exception as errormessage:
            print("Error:")
            print(str(errormessage))
            shutil.rmtree(temp_folder_base)
            shutil.rmtree(finished_folder)
    if one_saf:
        print("Finally creating the single saf for all items")
        try:
            if not project_name.isalnum():
                folder_name_clean = ""
                for char in project_name:
                    if not char.isalnum():
                        folder_name_clean = folder_name_clean + "_"
                    else:
                        folder_name_clean = folder_name_clean + char
                project_name = folder_name_clean
            shutil.make_archive(finished_folder / project_name,
                                format="zip", root_dir=temp_folder_base)
            shutil.rmtree(temp_folder_base)
        except Exception as errormessage:
            print("Error:")
            print(str(errormessage))
            shutil.rmtree(finished_folder)
    print("Finished creating multisaf")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('-f', help="Folder with saf data folders")
    parser.add_argument('-t', help="Spreadsheet with metadata")
    parser.add_argument('-onesaf', help="Create only one saf file", action='store_true')
    parser.add_argument(
        '-r', help="Folder with finished SAFs, will be placed next to saf data folder")
    args = parser.parse_args()
    if not args.t:
        print("ERROR: argument 'Spreadsheet with metadata' (-t) missing")
    else:
        if not args.f:
            print("ERROR: argument 'Folder with saf data folders' (-f) missing")
        else:
            if not args.r:
                print(
                    "ERROR: argument 'Folder with finished SAFs, will be placed"\
                        "next to saf data folder' (-r) missing")
            else:
                print("Creating multisaf with the following parameters:")
                print("-f, Folder with saf data folders: " + str(args.t))
                print("-t, Spreadsheet with metadata:" + str(args.f))
                print("-r, Folder with finished SAFs, will be placed next to saf data folder:"\
                    + str(args.r))
                print("-onesaf, create only one saf file: " + str(args.onesaf))
                create_saf(args.t, args.f, args.onesaf, args.r)
